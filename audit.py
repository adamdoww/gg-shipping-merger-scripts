#!/usr/bin/env python3
"""Audit a merged .xlsx against the source orders CSV.

Confirms that every cell from the original CSV is preserved exactly in the
output xlsx (modulo the inserted columns K + L, the appended notes column CD,
the totals row, and the documented J-header rename to "Shipping (customer paid)").

Usage:
    # Activate the virtualenv once per terminal session first:
    source .venv/bin/activate

    # Argument order is always ORDERS csv first, then the merged xlsx
    # (not the charges file). Use the same orders csv you fed to merge.py:
    python audit.py <orders.csv> <output.xlsx>

Exit status:
    0  — only expected differences found (xlsx is faithful to the source).
    1  — unexpected differences found (something changed that shouldn't have).
    2  — bad input (file missing, wrong shape, etc.).
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Headers whose values are money (compared as floats, rounded to 2 decimals).
MONEY_HEADERS = {'Subtotal', 'Shipping', 'Taxes', 'Total'}

# Differences we deliberately introduced and don't want flagged as failures.
EXPECTED_DIFFS = {
    # (xlsx_row, xlsx_col_letter): (csv_value, xlsx_value, reason)
    (1, 'J'): ('Shipping', 'Shipping (customer paid)',
               'header renamed for clarity per brief'),
}


def try_float(s):
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def csv_col_to_xlsx_col(csv_c: int) -> int:
    """Map a 0-based CSV column index to the matching 1-based xlsx column.

    K and L are inserted at xlsx positions 11 and 12; everything from old col K
    (csv index 10) onward shifts right by 2. The appended notes column at CD
    has no CSV counterpart.
    """
    return csv_c + 1 if csv_c <= 9 else csv_c + 3


def load_csv(path: Path) -> list[list[str]]:
    with open(path, newline='', encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    if not rows:
        raise SystemExit(f'CSV is empty: {path}')
    return rows


def diff(csv_rows: list[list[str]], xlsx_path: Path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header = csv_rows[0]
    money_csv_cols = {i for i, h in enumerate(header) if h.strip() in MONEY_HEADERS}

    n_csv_data = len(csv_rows) - 1
    n_xlsx_data = ws.max_row - 2  # header + totals
    if n_xlsx_data != n_csv_data:
        raise SystemExit(
            f'Row-count mismatch: CSV has {n_csv_data} data rows, '
            f'xlsx has {n_xlsx_data}. Aborting.'
        )

    # Stream all xlsx values into a dict for random access (read_only sheets
    # don't support .cell() well for repeated reads).
    xlsx_cells: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is not None and cell.value != '':
                xlsx_cells[(cell.row, cell.column)] = cell.value

    mismatches = []
    expected_hits = []
    total = 0
    for r in range(len(csv_rows)):
        csv_row = csv_rows[r]
        xlsx_row = r + 1
        for csv_c in range(len(csv_row)):
            total += 1
            xlsx_c = csv_col_to_xlsx_col(csv_c)
            csv_val = csv_row[csv_c]
            xlsx_val = xlsx_cells.get((xlsx_row, xlsx_c))

            if csv_c in money_csv_cols and r > 0:
                cf = try_float(csv_val)
                csv_norm = None if csv_val == '' else (round(cf, 2) if cf is not None else csv_val)
                xlsx_norm = (
                    round(xlsx_val, 2) if isinstance(xlsx_val, float)
                    else (None if xlsx_val in ('', None) else xlsx_val)
                )
                kind = 'money'
            else:
                csv_norm = None if csv_val == '' else csv_val
                xlsx_norm = None if xlsx_val in ('', None) else xlsx_val
                kind = 'text'

            if csv_norm == xlsx_norm:
                continue

            key = (xlsx_row, get_column_letter(xlsx_c))
            if key in EXPECTED_DIFFS:
                exp_csv, exp_xlsx, _reason = EXPECTED_DIFFS[key]
                if str(csv_val) == exp_csv and str(xlsx_val) == exp_xlsx:
                    expected_hits.append(key)
                    continue

            mismatches.append((xlsx_row, xlsx_c, csv_val, xlsx_val, kind))

    return total, mismatches, expected_hits, n_csv_data


def main():
    ap = argparse.ArgumentParser(
        description='Audit a merged .xlsx against the source orders CSV.'
    )
    ap.add_argument('orders_csv', help='Path to the source orders CSV.')
    ap.add_argument('output_xlsx', help='Path to the merged .xlsx to audit.')
    args = ap.parse_args()

    csv_path = Path(args.orders_csv).expanduser().resolve()
    xlsx_path = Path(args.output_xlsx).expanduser().resolve()
    for p in (csv_path, xlsx_path):
        if not p.exists():
            print(f'File not found: {p}', file=sys.stderr)
            sys.exit(2)

    csv_rows = load_csv(csv_path)
    total, mismatches, expected_hits, n_csv_data = diff(csv_rows, xlsx_path)

    print(f'Source CSV:  {csv_path}')
    print(f'Audited xlsx: {xlsx_path}')
    print()
    print(f'  CSV data rows:       {n_csv_data}')
    print(f'  Cells compared:      {total}')
    print(f'  Expected differences found: {len(expected_hits)}')
    for key in expected_hits:
        exp_csv, exp_xlsx, reason = EXPECTED_DIFFS[key]
        coord = f'{key[1]}{key[0]}'
        print(f'    {coord}: {exp_csv!r} → {exp_xlsx!r}  ({reason})')

    print(f'  Unexpected differences:     {len(mismatches)}')

    if mismatches:
        by_col = Counter(m[1] for m in mismatches)
        print()
        print('  Unexpected diffs by xlsx column:')
        for col, n in by_col.most_common(10):
            csv_c_back = col - 1 if col <= 10 else col - 3
            hname = csv_rows[0][csv_c_back] if csv_c_back < len(csv_rows[0]) else '?'
            print(f'    {get_column_letter(col)} ({hname}): {n}')
        print()
        print('  First 15 unexpected diffs:')
        for m in mismatches[:15]:
            print(f'    row {m[0]} col {get_column_letter(m[1])}: '
                  f'CSV={m[2]!r}  XLSX={m[3]!r}  ({m[4]})')
        print()
        print('FAIL — unexpected differences found.')
        sys.exit(1)

    print()
    print('PASS — every source cell preserved exactly (modulo expected differences).')


if __name__ == '__main__':
    main()
