#!/usr/bin/env python3
"""Merge Shopify shipping charges into an orders export.

Reads an orders CSV and a charges CSV, sums shipping_fee charges per order,
and writes a polished .xlsx with two new columns inserted at K and L:
  - Company Shipping Cost (carrier-charged amount)
  - Shipping Difference  (Shipping - Company Shipping Cost)

Values land only on the first row of each multi-line-item order. Orders
absent from the charges file are left blank in both new columns.

Usage:
    # Activate the virtualenv once per terminal session first:
    source .venv/bin/activate

    # Argument order is always ORDERS first, then CHARGES:
    python merge.py path/to/orders.csv path/to/charges.csv
    python merge.py orders.csv charges.csv -o custom_output.xlsx
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import zipfile
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --- Constants ----------------------------------------------------------------

CURRENCY_FMT = '"$"#,##0.00;[Red]-"$"#,##0.00'

# Original orders column positions (0-based)
NAME_COL = 0       # A — order number ("#GG24474")
SHIPPING_COL = 9   # J — customer-paid shipping

NEW_HEADER_J = 'Shipping (customer paid)'
NEW_HEADER_K = 'Company Shipping Cost'
NEW_HEADER_L = 'Shipping Difference'
NOTES_HEADER = 'Company Shipping Cost Notes'
NOTE_TEXT = (
    'No matching shipping charge in the charges export — '
    'order likely not yet billed/shipped.'
)

# These get currency formatting in the output (matched by header name).
MONEY_HEADERS = {
    'Subtotal',
    NEW_HEADER_J,
    NEW_HEADER_K,
    NEW_HEADER_L,
    'Taxes',
    'Total',
}

# Restrained business palette
HEADER_FILL = PatternFill('solid', fgColor='1F3A5F')        # navy
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
NEW_COL_FILL = PatternFill('solid', fgColor='FFF4D6')       # warm highlight
ZEBRA_FILL = PatternFill('solid', fgColor='F4F6FA')         # very light gray-blue
NEW_COL_ZEBRA = PatternFill('solid', fgColor='F5E9C2')      # darker shade of highlight
TOTALS_FILL = PatternFill('solid', fgColor='E3E8F0')
TOTALS_BORDER = Border(top=Side(style='medium', color='1F3A5F'))
POS_FILL = PatternFill('solid', fgColor='D9EAD3')
NEG_FILL = PatternFill('solid', fgColor='F4CCCC')
POS_FONT = Font(color='1A6A2E', bold=True)
NEG_FONT = Font(color='9C2A2A', bold=True)


# --- IO helpers ---------------------------------------------------------------

def parse_money(value) -> float | None:
    """Parse a money string to float; return None for blank/unparseable."""
    if value is None:
        return None
    s = str(value).strip().replace('$', '').replace(',', '')
    if s == '' or s == '""':
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_shipping_costs(charges_path: Path) -> dict[str, float]:
    """Sum shipping_fee Amounts per Order from the charges CSV."""
    df = pd.read_csv(charges_path, dtype=str, keep_default_na=False)
    required = {'Charge category', 'Order', 'Amount'}
    missing = required - set(df.columns)
    if missing:
        raise SystemExit(f'Charges CSV is missing required columns: {sorted(missing)}')

    df = df[df['Charge category'].str.strip() == 'shipping_fee']
    df = df[df['Order'].str.strip() != '']
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0.0)
    grouped = df.groupby(df['Order'].str.strip())['Amount'].sum().round(2)
    return grouped.to_dict()


def load_orders(orders_path: Path) -> tuple[list[str], list[list[str]]]:
    """Read orders CSV as raw strings, preserving every field as-is."""
    with open(orders_path, newline='', encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    if not rows:
        raise SystemExit('Orders CSV is empty.')
    return rows[0], rows[1:]


# --- Build workbook -----------------------------------------------------------

def build_workbook(
    headers: list[str],
    data_rows: list[list[str]],
    shipping_costs: dict[str, float],
) -> tuple[Workbook, list[str], dict[str, int], int, int]:
    """Build the workbook with the two new columns inserted at K and L.

    Returns: workbook, new_headers, money_col_lookup (header -> 1-based col),
             last_data_row (1-based), totals_row_idx (1-based).
    """
    if headers[NAME_COL].strip() != 'Name':
        print(f'WARN: column A is "{headers[NAME_COL]}", expected "Name".', file=sys.stderr)
    if headers[SHIPPING_COL].strip() != 'Shipping':
        print(f'WARN: column J is "{headers[SHIPPING_COL]}", expected "Shipping".', file=sys.stderr)
    if len(headers) != 79:
        print(f'WARN: orders header has {len(headers)} columns, brief expects 79.', file=sys.stderr)

    # Build new header row: rename J, insert two new headers after J, append notes at the end.
    new_headers = list(headers)
    new_headers[SHIPPING_COL] = NEW_HEADER_J
    insert_at = SHIPPING_COL + 1
    new_headers = (
        new_headers[:insert_at]
        + [NEW_HEADER_K, NEW_HEADER_L]
        + new_headers[insert_at:]
        + [NOTES_HEADER]
    )

    # 1-based positions for money columns in the new layout
    money_col_lookup: dict[str, int] = {}
    for i, h in enumerate(new_headers):
        if h.strip() in MONEY_HEADERS:
            money_col_lookup[h.strip()] = i + 1
    notes_col = len(new_headers)  # CD, 1-based

    wb = Workbook()
    ws = wb.active
    ws.title = 'Orders + Shipping'
    ws.append(new_headers)

    processed_first_row: set[str] = set()
    matched_orders = 0

    for src in data_rows:
        # Pad short rows to header width.
        if len(src) < len(headers):
            src = list(src) + [''] * (len(headers) - len(src))

        order_name = src[NAME_COL].strip()
        shipping_raw = src[SHIPPING_COL].strip() if SHIPPING_COL < len(src) else ''

        # First row of an order = first time we see this order *with* a populated Shipping cell.
        is_first_row = (
            bool(order_name)
            and order_name not in processed_first_row
            and shipping_raw != ''
        )

        company_cost: float | str = ''
        diff: float | str = ''
        note: str = ''
        if is_first_row:
            processed_first_row.add(order_name)
            if order_name in shipping_costs:
                company_cost = shipping_costs[order_name]
                shipping_paid = parse_money(shipping_raw)
                if shipping_paid is not None:
                    diff = round(shipping_paid - company_cost, 2)
                matched_orders += 1
            else:
                note = NOTE_TEXT

        new_row = src[:insert_at] + [company_cost, diff] + src[insert_at:] + [note]

        # Convert money column strings to floats so Excel sees them as numbers.
        for header_name, col_1based in money_col_lookup.items():
            val = new_row[col_1based - 1]
            if isinstance(val, (int, float)) or val == '' or val is None:
                continue
            parsed = parse_money(val)
            if parsed is not None:
                new_row[col_1based - 1] = parsed

        ws.append(new_row)

    last_data_row = ws.max_row
    totals_row_idx = last_data_row + 1

    # Pre-compute totals so we can both write SUM formulas AND cache their results.
    totals_cache: dict[str, float] = {}
    for header_name in (NEW_HEADER_J, NEW_HEADER_K, NEW_HEADER_L):
        col_1based = money_col_lookup[header_name]
        s = 0.0
        for r in range(2, last_data_row + 1):
            v = ws.cell(row=r, column=col_1based).value
            if isinstance(v, (int, float)):
                s += v
        totals_cache[header_name] = round(s, 2)

    # Totals row with live SUM formulas (cached values get injected after save).
    totals = [''] * len(new_headers)
    totals[0] = 'TOTALS'
    for header_name in (NEW_HEADER_J, NEW_HEADER_K, NEW_HEADER_L):
        col_1based = money_col_lookup[header_name]
        letter = get_column_letter(col_1based)
        totals[col_1based - 1] = f'=SUM({letter}2:{letter}{last_data_row})'
    ws.append(totals)

    return wb, new_headers, money_col_lookup, last_data_row, totals_row_idx, matched_orders, totals_cache


# --- Styling ------------------------------------------------------------------

def style_workbook(
    ws,
    new_headers: list[str],
    money_col_lookup: dict[str, int],
    last_data_row: int,
    totals_row_idx: int,
) -> None:
    n_cols = len(new_headers)
    k_col = money_col_lookup[NEW_HEADER_K]
    l_col = money_col_lookup[NEW_HEADER_L]
    notes_col = n_cols  # CD — always last

    # Header row
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

    # Currency format for money columns (including totals row)
    for col_1based in money_col_lookup.values():
        for r in range(2, totals_row_idx + 1):
            ws.cell(row=r, column=col_1based).number_format = CURRENCY_FMT

    # Zebra striping + highlight K/L/Notes in a single pass over data rows
    notes_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    notes_font = Font(italic=True)
    for r in range(2, last_data_row + 1):
        is_even = (r % 2) == 0
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if c == k_col or c == l_col or c == notes_col:
                cell.fill = NEW_COL_ZEBRA if is_even else NEW_COL_FILL
            elif is_even:
                cell.fill = ZEBRA_FILL
        notes_cell = ws.cell(row=r, column=notes_col)
        notes_cell.alignment = notes_alignment
        notes_cell.font = notes_font

    # Conditional formatting on Shipping Difference (data range + totals row)
    diff_range = f'{get_column_letter(l_col)}2:{get_column_letter(l_col)}{totals_row_idx}'
    ws.conditional_formatting.add(
        diff_range,
        CellIsRule(operator='greaterThan', formula=['0'], fill=POS_FILL, font=POS_FONT),
    )
    ws.conditional_formatting.add(
        diff_range,
        CellIsRule(operator='lessThan', formula=['0'], fill=NEG_FILL, font=NEG_FONT),
    )

    # Totals row styling
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=totals_row_idx, column=c)
        cell.font = Font(bold=True)
        cell.fill = TOTALS_FILL
        cell.border = TOTALS_BORDER
    ws.row_dimensions[totals_row_idx].height = 20

    # Column widths: auto-fit with min/max caps (notes column gets a fixed 45)
    MAX_WIDTH = 38
    MIN_WIDTH = 10
    NOTES_WIDTH = 45
    for c in range(1, n_cols + 1):
        if c == notes_col:
            ws.column_dimensions[get_column_letter(c)].width = NOTES_WIDTH
            continue
        max_len = len(str(new_headers[c - 1]))
        for r in range(2, last_data_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or v == '':
                continue
            if isinstance(v, float):
                length = len(f'{v:,.2f}') + 1  # leave room for $
            else:
                length = len(str(v))
            if length > max_len:
                max_len = length
            if max_len >= MAX_WIDTH:
                break
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, MIN_WIDTH), MAX_WIDTH)

    # Autofilter on header row across full data range incl. notes column (exclude totals row)
    ws.auto_filter.ref = f'A1:{get_column_letter(n_cols)}{last_data_row}'


# --- Post-save: cache formula results + error check ---------------------------

def _inject_cached_value(xml_text: str, cell_ref: str, value: float) -> tuple[str, int]:
    """Populate the cached-value <v> tag for a formula cell at the given A1 ref.

    openpyxl already emits an empty <v /> after each formula's <f>; we replace
    that with <v>{value}</v>. If the cell already has a populated <v>, we
    overwrite it. If somehow there's no <v> at all, we insert one.
    """
    formatted = f'{value:.2f}'
    # Case 1: empty self-closing <v />  →  <v>value</v>
    pat_empty = re.compile(
        rf'(<c\s[^>]*r="{re.escape(cell_ref)}"[^>]*>'
        rf'<f(?:\s[^>]*)?>[^<]*</f>)'
        rf'<v\s*/>'
        rf'(</c>)'
    )
    new_text, n = pat_empty.subn(rf'\1<v>{formatted}</v>\2', xml_text)
    if n:
        return new_text, n
    # Case 2: existing populated <v>...</v>  →  overwrite
    pat_full = re.compile(
        rf'(<c\s[^>]*r="{re.escape(cell_ref)}"[^>]*>'
        rf'<f(?:\s[^>]*)?>[^<]*</f>)'
        rf'<v>[^<]*</v>'
        rf'(</c>)'
    )
    new_text, n = pat_full.subn(rf'\1<v>{formatted}</v>\2', xml_text)
    if n:
        return new_text, n
    # Case 3: no <v> at all  →  insert one before </c>
    pat_none = re.compile(
        rf'(<c\s[^>]*r="{re.escape(cell_ref)}"[^>]*>'
        rf'<f(?:\s[^>]*)?>[^<]*</f>)'
        rf'(</c>)'
    )
    new_text, n = pat_none.subn(rf'\1<v>{formatted}</v>\2', xml_text)
    return new_text, n


def cache_totals_in_xlsx(
    xlsx_path: Path,
    totals_row_idx: int,
    money_col_lookup: dict[str, int],
    totals_cache: dict[str, float],
) -> None:
    """Post-process the .xlsx to add cached <v> results next to each totals formula.

    Without this, some viewers (Numbers, Google Sheets, older Excel) render the
    totals cells as blank until they recompute. With it, the value displays
    immediately while the live SUM formula is preserved.
    """
    sheet_path = 'xl/worksheets/sheet1.xml'
    tmp_path = xlsx_path.with_suffix(xlsx_path.suffix + '.tmp')
    injected_refs: list[str] = []
    with zipfile.ZipFile(xlsx_path, 'r') as zin, \
            zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                xml_text = data.decode('utf-8')
                for header in (NEW_HEADER_J, NEW_HEADER_K, NEW_HEADER_L):
                    col = money_col_lookup[header]
                    ref = f'{get_column_letter(col)}{totals_row_idx}'
                    xml_text, n = _inject_cached_value(xml_text, ref, totals_cache[header])
                    if n == 0:
                        raise RuntimeError(
                            f'Could not inject cached value for totals cell {ref}'
                        )
                    injected_refs.append(ref)
                data = xml_text.encode('utf-8')
            zout.writestr(item, data)
    tmp_path.replace(xlsx_path)


def check_formula_errors(xlsx_path: Path, totals_row_idx: int) -> tuple[int, int, int]:
    """Re-open the workbook and check formula health.

    Returns (formula_count, error_count, cached_totals_count).
    cached_totals_count = how many totals-row cells have a numeric cached value.
    """
    wb_f = load_workbook(xlsx_path, data_only=False)
    ws_f = wb_f.active
    formula_count = 0
    error_count = 0
    for row in ws_f.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formula_count += 1
            elif cell.data_type == 'e':
                error_count += 1
                print(f'  formula ERROR at {cell.coordinate}: {cell.value}', file=sys.stderr)

    wb_v = load_workbook(xlsx_path, data_only=True)
    ws_v = wb_v.active
    cached_totals_count = 0
    for cell in ws_v[totals_row_idx]:
        if isinstance(cell.value, (int, float)):
            cached_totals_count += 1
    return formula_count, error_count, cached_totals_count


def next_available_path(path: Path) -> Path:
    """If `path` exists, return path with `_2`, `_3`, ... appended to the stem
    until an unused name is found."""
    if not path.exists():
        return path
    stem, suffix, parent = path.stem, path.suffix, path.parent
    n = 2
    while True:
        candidate = parent / f'{stem}_{n}{suffix}'
        if not candidate.exists():
            return candidate
        n += 1


# --- Core entry point (shared by CLI and web app) -----------------------------

def merge_orders_and_charges(
    orders_path: Path,
    charges_path: Path,
    output_path: Path,
) -> dict:
    """Run the full merge and write the styled .xlsx to ``output_path``.

    This is the single source of truth for the merge logic, called by both the
    CLI (``main``) and the Streamlit web app. Returns a dict of summary stats.
    """
    shipping_costs = load_shipping_costs(charges_path)
    headers, data_rows = load_orders(orders_path)
    wb, new_headers, money_col_lookup, last_data_row, totals_row_idx, matched_orders, totals_cache = \
        build_workbook(headers, data_rows, shipping_costs)
    style_workbook(wb.active, new_headers, money_col_lookup, last_data_row, totals_row_idx)

    # Tell Excel to force a full recalc when the file is opened (belt + suspenders
    # alongside the cached <v> values we inject below).
    wb.calculation.fullCalcOnLoad = True

    wb.save(output_path)

    # Post-save: inject cached results next to each SUM formula so viewers display
    # the totals immediately instead of showing blank until they recompute.
    cache_totals_in_xlsx(output_path, totals_row_idx, money_col_lookup, totals_cache)

    formula_count, error_count, cached_totals = check_formula_errors(output_path, totals_row_idx)

    notes_col = len(new_headers)
    unbilled_count = sum(
        1 for r in range(2, last_data_row + 1)
        if wb.active.cell(row=r, column=notes_col).value == NOTE_TEXT
    )

    return {
        'output_path': output_path,
        'num_columns': len(new_headers),
        'last_column_letter': get_column_letter(len(new_headers)),
        'data_rows': last_data_row - 1,
        'matched_orders': matched_orders,
        'unique_orders_in_charges': len(shipping_costs),
        'unbilled_count': unbilled_count,
        'total_shipping_paid': totals_cache[NEW_HEADER_J],
        'total_company_cost': totals_cache[NEW_HEADER_K],
        'total_difference': totals_cache[NEW_HEADER_L],
        'formula_count': formula_count,
        'error_count': error_count,
        'cached_totals': cached_totals,
    }


# --- Main ---------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(
        description='Merge Shopify shipping charges into an orders export.',
    )
    ap.add_argument('orders', help='Path to orders_export.csv')
    ap.add_argument('charges', help='Path to charges_export.csv')
    ap.add_argument(
        '-o', '--output',
        help='Output .xlsx path. Default: '
             '<project>/Merged-CSVs/orders_with_shipping_cost_<date>.xlsx',
    )
    ap.add_argument(
        '-f', '--force', action='store_true',
        help='Overwrite the output file if it already exists. '
             'By default the script appends _2, _3, etc. to avoid clobbering prior runs.',
    )
    args = ap.parse_args()

    orders_path = Path(args.orders).expanduser().resolve()
    charges_path = Path(args.charges).expanduser().resolve()
    for p in (orders_path, charges_path):
        if not p.exists():
            raise SystemExit(f'File not found: {p}')

    if args.output:
        output_path = Path(args.output).expanduser().resolve()
    else:
        project_root = Path(__file__).resolve().parent
        output_dir = project_root / 'Merged-CSVs'
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / f'orders_with_shipping_cost_{date.today().isoformat()}.xlsx'

    if output_path.exists() and not args.force:
        output_path = next_available_path(output_path)

    stats = merge_orders_and_charges(orders_path, charges_path, output_path)

    print(f'Wrote {stats["output_path"]}')
    print(f'  Columns:                       {stats["num_columns"]} '
          f'(A..{stats["last_column_letter"]})')
    print(f'  Data rows:                     {stats["data_rows"]}')
    print(f'  Orders matched to charges:     {stats["matched_orders"]}')
    print(f'  Unique orders in charges file: {stats["unique_orders_in_charges"]}')
    print(f'  Unbilled orders (notes added): {stats["unbilled_count"]}')
    print(f'  Totals (cached values):        '
          f'J=${stats["total_shipping_paid"]:.2f}  '
          f'K=${stats["total_company_cost"]:.2f}  '
          f'L=${stats["total_difference"]:.2f}')
    print(f'  Formula cells: {stats["formula_count"]}   Errors: {stats["error_count"]}   '
          f'Cached totals-row values: {stats["cached_totals"]}/3')


if __name__ == '__main__':
    main()
